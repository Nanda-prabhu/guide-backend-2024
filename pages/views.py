
import os
from .models import Team, Credit
from openpyxl import Workbook
from random import randrange
from django.core.mail import send_mail
from django.http import HttpResponse, HttpResponseRedirect
from guide_project.settings import EMAIL_HOST_USER
from django.contrib.auth.models import User
from django.shortcuts import redirect, render
from django.contrib.auth.models import User, auth
from django.contrib import messages
from pages.models import Guide, Team, Otp, Otp_Two, Temp_Team
from accounts.models import BestTeam
from guide_project.storages_backends import MediaStorage

# Create your views here.


def home(request):
    user = request.user
    is_guide = False
    if user.is_authenticated:
        guide = Guide.objects.filter(email=user.email)
        if guide:
            is_guide = True

    return render(request, 'Home/home.html', context={'is_guide': is_guide})


def no_of_stud(request):

    return render(request, 'no_of_stud/no_of_stud.html')


def guides(request):
    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        emp_id = request.POST['emp_id']
        serial_no = request.POST['serial_no']
        designation = request.POST['designation']
        domain_1 = request.POST['domain_1']
        domain_2 = request.POST['domain_2']
        domain_3 = request.POST['domain_3']
        email = request.POST['email']
        myImage = request.FILES['myImage']

        name = first_name + ' ' + last_name

        serial_no = int(serial_no)

    #    if serial_no >= 1 and serial_no <= 52:
    #         vacancy = 7
    #     elif serial_no >= 53 and serial_no <= 79:
    #         vacancy = 4
    #     else:
    #         vacancy = 3

        if Guide.objects.filter(serial_no=serial_no).exists():
            messages.error(
                request, 'This serial number already exists. Please enter your own serial number')
            return redirect('guides')
        else:
            guide = Guide(serial_no=serial_no, emp_id=emp_id, designation=designation, name=name, domain_1=domain_1, domain_2=domain_2,
                          domain_3=domain_3, email=email, myImage=myImage)

            guide.save()
        return redirect('login')
    else:

        return render(request, 'adminregister/aform.html')


def submitted(request):
    auth.logout(request)
    return render(request, 'submitted.html')


def mail1(request):
    user = request.user
    if user.is_authenticated:
        if request.method == 'POST':
            # email_1 = request.POST['email_1']
            email_2 = request.POST['email_2']

            no = randrange(1000, 9999)
            # print("2nd MEMBER OTP IS: ", no)
            if Otp_Two.objects.filter(user_email=email_2).exists():
                t = Otp_Two.objects.filter(user_email=email_2)
                t.delete()
                print("OTP DELETED AND SENT AGAIN")
            if User.objects.filter(email=email_2).exists():
                messages.error(
                    request, 'The mail id already registered!')
                return redirect('mail1')
            elif Team.objects.filter(student_2_email=email_2).exists():
                messages.error(
                    request, 'The Second mail id already exists with another team!')
                return redirect('mail1')
            elif Team.objects.filter(student_1_email=email_2).exists():
                messages.error(
                    request, 'The Second mail id already exists with another team!')
                return redirect('mail1')
            email = Otp_Two.objects.create(
                user_email=email_2, temp_email=request.user.email, otp=no)
            email.save()
            send_mail(
                'YOUR OTP for verification',
                'Your OTP is: {}'.format(no),
                EMAIL_HOST_USER,
                [email_2, ],
                fail_silently=False,
            )

            return redirect('verify1')
        else:
            user = request.user
            context = {
                'user': user,
            }
            return render(request, 'Register/mail1.html', context)
    else:
        messages.error(request, "You're not logged In!")
        return redirect('login')


def verify1(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            otp = request.POST['otp']
            if Otp_Two.objects.filter(temp_email=request.user.email).exists():
                g_otp = Otp_Two.objects.filter(
                    temp_email=request.user.email).get()
            else:
                return redirect('mail1')

            if str(otp)== str(g_otp.otp):
                return redirect('project-details-2')
            else:
                messages.error(request, "Incorrect OTP try again")
                return redirect('mail1')
        else:
            return render(request, 'Register/verify1.html')
    else:
        messages.error(request, "You're not logged In!")
        return redirect('login')


def project_details_1(request):
    guides = Guide.objects.order_by('serial_no')
    curr_user = request.user
    if curr_user.is_authenticated:
        if Team.objects.filter(teamID=curr_user.username).exists():
            is_team = Team.objects.filter(teamID=curr_user.username).get()
            guide_inst = Guide.objects.filter(serial_no=is_team.guide)
            guide_inst.vacancy += 1
            guide_inst.save()
            is_team.delete()
            is_user = User.objects.filter(username=is_team.teamID)
            is_user.delete()
            messages.info(
                request, 'Your team is removed please to the process again!!')
            return render(request, 'Register/register.html')

        if request.method == 'POST':

            project_name = request.POST['project_name']
            project_domain = request.POST['project_domain']
            project_description = request.POST['project_description']
            reg_no_1 = request.POST['reg_no_1']
            if len(reg_no_1) > 8:
                messages.error(request, 'Register Number be 8 digits long.')
                return redirect('project-details-1')
            student_1_no = request.POST['student_1_no']
            if len(student_1_no) > 10:
                messages.error(request, 'Number must of 10 digits.')
                return redirect('project-details-1')

            student_1_name = curr_user.first_name + ' ' + curr_user.last_name
            student_1_email = curr_user.email

            user = User.objects.get(username=curr_user.username)
            # print("TYPE OF user.id: ", type(user.id))

            if Temp_Team.objects.filter(student_1_email=curr_user.email).exists():
                obj = Temp_Team.objects.filter(
                    student_1_email=curr_user.email).get()
                obj.delete()

            if Temp_Team.objects.filter(reg_no_1=reg_no_1).exists():
                messages.error(
                    request, 'The Register Number already exists in another team.')
                return redirect('project-details-1')

            temp_team = Temp_Team.objects.create(
                project_name=project_name,
                project_domain=project_domain,
                project_description=project_description,
                no_of_members='1', reg_no_1=reg_no_1,
                student_1_name=student_1_name,
                student_1_email=student_1_email,
                student_1_no=student_1_no
            )
            # print('temp_team: ', temp_team.project_name)
            temp_team.save()

            context = {
                'user': curr_user,
                'guides': guides,
            }

            return render(request, 'GuideList/guide.html', context)
        else:
            context = {
                'user': curr_user
            }
            return render(request, '1_project_form/1_project_form.html', context)
    else:
        messages.error(request, "You're not logged In!")
        return redirect('login')


def project_details_2(request):
    curr_user = request.user
    guides = Guide.objects.order_by('serial_no')
    if curr_user.is_authenticated:
        student_2_email = Otp_Two.objects.filter(
            temp_email=curr_user.email).first()

        if Team.objects.filter(teamID=curr_user.username).exists():
            is_team = Team.objects.filter(teamID=curr_user.username).get()
            is_team.delete()
            is_user = User.objects.filter(username=is_team.teamID)
            is_user.delete()
            messages.info(
                request, 'Your team is removed please do the process again!!')
            return render(request, 'Login/login.html')
        if request.method == 'POST':

            project_name = request.POST['project_name']
            project_domain = request.POST['project_domain']
            project_description = request.POST['project_description']
            reg_no_1 = request.POST['reg_no_1']
            student_1_no = request.POST['student_1_no']
            if len(reg_no_1) > 8:
                messages.error(request, 'Register Number must be 8 digits long.')
                return redirect('project-details-1')
            student_1_no = request.POST['student_1_no']
            if len(student_1_no) > 10:
                messages.error(request, 'Number must contain 10 digits.')
                return redirect('project-details-2')

            student_1_name = curr_user.first_name + ' ' + curr_user.last_name
            student_1_email = curr_user.email

            first_name_2 = request.POST['first_name_2']
            last_name_2 = request.POST['last_name_2']
            reg_no_2 = request.POST['reg_no_2']
            student_2_no = request.POST['student_2_no']
            if len(reg_no_2) > 8:
                messages.error(request, 'Register Number must be 8 digits long.')
                return redirect('project-details-2')
            student_2_no = request.POST['student_2_no']
            if len(student_2_no) > 10:
                messages.error(request, 'Number must contain 10 digits.')
                return redirect('project-details-2')

            student_2_name = first_name_2 + ' ' + last_name_2

            user = User.objects.get(username=curr_user.username)
            # print("TYPE OF user.id: ", type(user.id))

            if Team.objects.filter(reg_no_1=reg_no_1).exists():
                messages.error(
                    request, '1st Register Number already exists in another team.')
                return redirect('project-details-2')
            elif Team.objects.filter(reg_no_2=reg_no_2).exists():
                messages.error(
                    request, '2nd Register Number already exists in another team.')
                return redirect('project-details-2')

            if Team.objects.filter(student_1_no=student_1_no).exists():
                messages.error(
                    request, '1st Phone Number already exists in another team.')
                return redirect('project-details-2')
            elif Team.objects.filter(student_2_no=student_2_no).exists():
                messages.error(
                    request, '2nd Phone Number already exists in another team.')
                return redirect('project-details-2')

            if Temp_Team.objects.filter(student_1_email=curr_user.email).exists():
                    obj = Temp_Team.objects.filter(
                        student_1_email=curr_user.email).get()
                    obj.delete()

            temp_team = Temp_Team.objects.create(project_name=project_name, project_domain=project_domain, project_description=project_description, no_of_members='2', reg_no_1=reg_no_1,
                                                student_1_name=student_1_name, student_1_email=student_1_email, student_1_no=student_1_no, reg_no_2=reg_no_2,  student_2_name=student_2_name, student_2_email=student_2_email.user_email, student_2_no=student_2_no)

            temp_team.save()

            context = {
                'guides': guides,
            }

            return render(request, 'GuideList/guide.html', context)
        else:
            print('INSIDE GET REQUEST ELSE')
            context = {
                'email': student_2_email,
                'user': curr_user,
            }
            return render(request, '2_project_form/2_project_form.html', context)
    else:
        messages.error(request, "You're not logged In!")
        return redirect('login')    


def select_guide(request):

    if not request.user.is_authenticated:
        messages.error(request, "You're not logged In!")
        return redirect('login')

    guides = Guide.objects.order_by('serial_no')  # object
    if request.method == 'POST':

        return redirect('guide-selected')

    context = {
        'guides': guides,
    }

    return render(request, 'GuideList/guide.html', context)


def temp_team_2(request):
    user = request.user
    if not user.is_authenticated:
        messages.error(request, "You're not logged In!")
        return redirect('login')
    if request.method == 'POST':
        if Temp_Team.objects.filter(student_1_email=user.email).exists():
            temp_team = Temp_Team.objects.filter(
                student_1_email=user.email).get()
            project_name = request.POST['project_name']
            project_domain = request.POST['project_domain']
            project_description = request.POST['project_description']
            reg_no_1 = request.POST['reg_no_1']
            if len(reg_no_1) > 8:
                messages.error(request, 'Register Number be 8 digits long.')
                return redirect('project-details-1')
            student_1_no = request.POST['student_1_no']
            if len(student_1_no) > 10:
                messages.error(request, 'Number must of 10 digits.')
                return redirect('project-details-1')

            guide = request.POST['guide']
            guide_email = request.POST['guide_email']

            student_1_name = user.first_name + ' ' + user.last_name
            student_1_email = user.email

            reg_no_2 = request.POST['reg_no_2']
            if len(reg_no_2) > 8:
                messages.error(request, 'Register Number be 8 digits long.')
                return redirect('project-details-1')
            student_2_no = request.POST['student_2_no']
            if len(student_2_no) > 10:
                messages.error(request, 'Number must of 10 digits.')
                return redirect('project-details-1')

            guide = request.POST['guide']
            guide_email = request.POST['guide_email']

            student_2_name = user.first_name + ' ' + user.last_name

            if Otp_Two.objects.filter(temp_email=user.email).exists():
                student_2_email = Otp_Two.objects.filter(
                    temp_email=user.email).first()

                student_2_email = student_2_email.user_email

            temp_team.delete()

            temp_team = Temp_Team.objects.create(project_name=project_name, project_domain=project_domain, project_description=project_description, student_1_name=student_1_name, student_1_email=student_1_email,
                                                 reg_no_1=reg_no_1, student_1_no=student_1_no, student_2_name=student_2_name, student_2_email=student_2_email, reg_no_2=reg_no_2, no_of_members='2', guide=guide, guide_email=guide_email)

            if Guide.objects.filter(email=guide_email).exists():
                guide = Guide.objects.filter(email=guide_email).get()
                if guide.vacancy == 0:
                    redirect('select-guide')
                temp_team.save()
                context = {
                    'id': guide.serial_no,
                    'team': temp_team,
                    'user': user
                }

                if temp_team.no_of_members == '2':
                    return render(request, 'confirmation_2/confirmation.html', context)
                else:
                    return render(request, 'confirmation_1/confirmation.html', context)
            else:
                temp_team.save()
                messages.info(
                    request, 'You have not selected the guide please select!')
                return redirect('select-guide')
        else:
            return redirect('custom_page_not_found_view')


def temp_team_1(request):
    user = request.user
    if not user.is_authenticated:
        messages.error(request, "You're not logged In!")
        return redirect('login')
    if request.method == 'POST':
        if Temp_Team.objects.filter(student_1_email=user.email).exists():
            temp_team = Temp_Team.objects.filter(
                student_1_email=user.email).get()
            project_name = request.POST['project_name']
            project_domain = request.POST['project_domain']
            project_description = request.POST['project_description']
            reg_no_1 = request.POST['reg_no_1']
            if len(reg_no_1) > 8:
                messages.error(request, 'Register Number be 8 digits long.')
                return redirect('project-details-1')
            student_1_no = request.POST['student_1_no']
            if len(student_1_no) > 10:
                messages.error(request, 'Number must of 10 digits.')
                return redirect('project-details-1')

            student_1_name = user.first_name + ' ' + user.last_name
            student_1_email = user.email

            guide_inst = Guide.objects.filter(
                email=temp_team.guide_email).get()
            guide, guide_email = guide_inst.name, guide_inst.email

            temp_team.delete()

            temp_team = Temp_Team.objects.create(project_name=project_name, project_domain=project_domain, project_description=project_description, student_1_name=student_1_name,
                                                 student_1_email=student_1_email, reg_no_1=reg_no_1, student_1_no=student_1_no, no_of_members='1', guide=guide, guide_email=guide_email)

            if Guide.objects.filter(email=guide_email).exists():
                guide = Guide.objects.filter(email=guide_email).get()
                if guide.vacancy == 0:
                    redirect('select-guide')
                temp_team.save()
                context = {
                    'id': guide.serial_no,
                    'team': temp_team,
                    'user': user
                }

                if temp_team.no_of_members == '2':
                    return render(request, 'confirmation_2/confirmation.html', context)
                else:
                    return render(request, 'confirmation_1/confirmation.html', context)
            else:
                temp_team.save()
                messages.info(
                    request, 'You have not selected the guide please select!')
                return redirect('select-guide')
        else:
            return redirect('custom_page_not_found_view')
    else:
        temp_team = Temp_Team.objects.filter(student_1_email=user.email).get()
        guide = Guide.objects.filter(email=temp_team.guide_email).get()
        context = {
            'id': guide.serial_no,
            'guide': guide,
            'team': temp_team,
            'user': user
        }
        if temp_team.no_of_members == '2':
            return render(request, 'temp_team_2/temp_team_2.html')
        else:
            return render(request, 'temp_team_1/temp_team_1.html')


# For confirmation page


def guide_selected(request, id):

    guide_inst = Guide.objects.get(serial_no=id)
    user = request.user
    if not user.is_authenticated:
        messages.error(request, "You're not logged In!")
        return ('login')
    # you can get teamID from username as both are same.
    try:
        temp_team = Temp_Team.objects.get(student_1_email=user.email)
    except:
        return redirect('/accounts/login')

    temp_team.guide = guide_inst.name
    temp_team.guide_email = guide_inst.email
    temp_team.save()

    obj = Otp_Two.objects.filter(temp_email=user.email)

    if request.method == 'POST':

        if Team.objects.filter(student_1_email=user.email).exists():
            team = Team.objects.get(student_1_email=user.email)
            team.guide = guide_inst.name
            team.guide_email = guide_inst.email
        else:
            team = Team.objects.create(project_name=temp_team.project_name, project_domain=temp_team.project_domain, project_description=temp_team.project_description, no_of_members=temp_team.no_of_members, reg_no_1=temp_team.reg_no_1, student_1_name=temp_team.student_1_name,
                                   student_1_email=temp_team.student_1_email, student_1_no=temp_team.student_1_no, reg_no_2=temp_team.reg_no_2, student_2_name=temp_team.student_2_name, student_2_email=temp_team.student_2_email, student_2_no=temp_team.student_2_no, guide=guide_inst.name, guide_email=guide_inst.email)

        new_username = "CSE-%03d" % (team.id)  # CSE-001, CSE-002, ....
        team.teamID = new_username
        user.username = new_username
        obj.delete()
        temp_team.delete()
        team.save()

        if team.no_of_members == '2':
            if guide_inst.vacancy == 0:
                messages.error(request, "Guide vacancy has become Zero(0)! Please chose another guide")
                return HttpResponseRedirect('/select-guide/')
                # return HttpResponse('0 Vacancies for the selected Guide.Kindly chose another')
            user.save()
            guide_inst.vacancy -= 1
            guide_inst.save()
            send_mail(
                'CONFIRMATION FOR FINAL YEAR PROJECT REGISTRATION',
                'Hi, Thank you for registering here is your details:' + '\n\nTeam ID: ' + team.teamID + '\n\nProject Name: ' + team.project_name + '\n\nProject Description: ' + team.project_description + '\n\nGuide Name: ' + guide_inst.name + '\n\nGuide Email: ' + guide_inst.email + '\n\nNo. of members: ' + team.no_of_members + '\n\nMembers: ' + team.student_1_name + ' and '+team.student_2_name +
                '\n\nNow you can login with your teamID and password(The one you created earlier)',
                EMAIL_HOST_USER,
                [user.email, team.student_2_email],
                fail_silently=False,
            )

        else:
            if guide_inst.vacancy == 0:
                messages.error(request, "Guide vacancy has become Zero(0)! Please chose another guide")
                return HttpResponseRedirect('/select-guide/')
                # return HttpResponse('0 Vacancies for the selected Guide.Kindly chose another')
            user.save()
            guide_inst.vacancy -= 1
            guide_inst.save()
            send_mail(
                'CONFIRMATION FOR FINAL YEAR PROJECT REGISTRATION',
                'Hi, Thank you for registering here is your details:' + '\n\nTeam ID: ' + team.teamID + '\n\nProject Name: ' + team.project_name + '\n\nProject Description: ' + team.project_description + '\n\nGuide Name: ' + guide_inst.name + '\n\nGuide Email: ' + guide_inst.email + '\n\nNo. of members: ' + team.no_of_members + '\n\nMembers: ' + team.student_1_name +
                '\n\nNow you can login with your TEAMID and password(The one you created earlier)',
                EMAIL_HOST_USER,
                [user.email, ],
                fail_silently=False,
            )
            team.save()
            temp_team.delete()
        return redirect('team-dashboard')
        # return render(request, 'submitted.html')
    context = {
        'guide': guide_inst,
        'team': temp_team,
        'id': id,
        'user': user,
    }
    # print("TEAM MEM: ", temp_team.no_of_members)
    if temp_team.no_of_members == '2':
        print('CONFIRM 2')
        return render(request, 'confirmation_2/confirmation.html', context)
    else:
        print('CONFIRM 1')
        return render(request, 'confirmation_1/confirmation.html', context)


def credits(request):
    credits = Credit.objects.all()
    context = {
        'credits': credits,
    }
    return render(request, 'credits/credit.html', context)


def search(request):
    if not request.user.is_authenticated:
        messages.error(request, "You're not logged In!")
        return redirect('login')
    print("INSIDE SEARCH FUNCTION: ")
    queryset_list = Guide.objects.order_by('serial_no')

    if 'name' in request.GET:
        print("INSIDE GET IF")
        name = request.GET['name']
        if name:
            print("INSIDE NAME IF..")
            queryset_list = queryset_list.filter(name__icontains=name)
            # print(queryset_list)
    context = {
        'guides': queryset_list,
    }

    return render(request, 'search.html', context)


def reset_password(request):
    if request.method == 'POST':
        email = request.POST['email']
        teamID = request.POST['teamID']
        password = request.POST['password']
        password1 = request.POST['password1']

        if password == password1:
            special_characters = "[~\!@#\$%\^&\*\(\)_\+{}\":;'\[\]]"
            if len(password) < 8:
                messages.error(
                    request, 'Password length must be atleast 8 character.')
                return redirect('reset-password')

            # Check for dig
            if not any(char.isdigit() for char in password):
                messages.error(
                    request, 'Password must contain at least 1 digit.')
                return redirect('reset-password')

            # Check for alpha
            if not any(char.isalpha() for char in password):
                messages.error(
                    request, 'Password must contain at least 1 letter and must be alpha-numeric.')
                return redirect('reset-password')

            # Check spl char
            if not any(char in special_characters for char in password):
                messages.error(
                    request, 'Password must contain at least 1 special character')
                return redirect('reset-password')

            # if Guide.objects.filter(emp_id=teamID).exists():
            #     if User.objects.filter(username=email).exists():
            #         pass
            # else:
            #     return redirect('login')

            # Check for user existence
            temp, id = teamID.split('-')
            # temp = int(teamID)
            # print('id is: ', id)
            if Guide.objects.filter(emp_id=id).exists():
                if User.objects.filter(username=email).exists():
                    user = User.objects.filter(username=email).get()
                    user.email = email
                    user.set_password(password)
                    user.save()
                    messages.success(request, 'Password Changed successfully!')
                    return redirect('login')
            elif User.objects.filter(username=teamID).exists():
                user = User.objects.filter(username=teamID).get()
                user.email = email
                user.set_password(password)
                user.save()
                messages.success(request, 'Password Changed successfully!')
                return redirect('login')
            else:
                messages.error(
                    request, 'Given Email-id/team ID/Emp-id does not exist.  Try Again!')
                return redirect('reset-password')
        else:
            messages.error(request, 'Password not matching!')
            return redirect('reset-password')

    return render(request, 'resetpass/resetpass.html')


def doc_upload(request):
    user = request.user
    if user.is_authenticated:
        if BestTeam.objects.filter(teamID=user.username).exists():
            if request.method == 'POST':
                print('INSIDE POST DOC UPLOAD')
                team = Team.objects.filter(teamID=user.username).get()

                media_storage = MediaStorage()
                # documents/CSE-001/filename.ppt
                file_path_bucket = 'documents/{0}/'.format(
                    request.user.username)
                # print('Path is: ', file_path_bucket)
                # Below line always returns False
                # print('Check Dir', media_storage.exists(file_path_bucket))
                if media_storage.exists(file_path_bucket):
                    media_storage.delete(file_path_bucket)

                app_video_directory_within_bucket = 'App_Based/{username}'.format(
                    username=request.user)

                product_video_directory_within_bucket = 'Product_Based/{username}'.format(
                    username=request.user)

                # UN-COMMENT THE BELOW ONCE FINISHED WITH THE DROP DOWN PART IN THE DOC UPLOAD PAGE (upload_docs/docs.html) TO MAKE THE CHANGES AFFECT IN THE BACKEND
                type = request.POST['type']
                # print('FILES value: ', request.FILES)

                if request.FILES:
                    # ppt
                    if request.FILES.get('ppt'):
                        ppt = request.FILES['ppt']
                        # ppt and docs max size 9MB each (9 MB = 94,37,184 B)
                        if ppt.size > 9437184:
                            messages.error(
                                request, "PPT size must be less than 9 MB")
                            return redirect('upload')
                        team.ppt = ppt

                    # document
                    if request.FILES.get('document'):
                        document = request.FILES['document']
                        if document.size > 9437184:
                            messages.error(
                                request, "Document size must be less than 9 MB")
                            return redirect('upload')
                        team.document = document

                    # Research Paper
                    if request.FILES.get('rs_paper'):
                        rs_paper = request.FILES['rs_paper']
                        # guide_form and rs_paper 500 kb each (Total 1 MB for pdfs) (1 MB = 10,48,576 B)
                        if rs_paper.size > 1048576:
                            messages.error(
                                request, "Research Paper size must be less than 500kb")
                            return redirect('upload')
                        team.rs_paper = rs_paper

                    # Guide Form
                    if request.FILES.get('guide_form'):
                        guide_form = request.FILES['guide_form']
                        if guide_form.size > 1048576:
                            messages.error(
                                request, "Guide Form size must be less than 500kb")
                            return redirect('upload')
                        team.guide_form = guide_form

                    # size is not set correctly pls change later
                    if request.POST.get('demo_video'):
                        if type == 'App Based':
                            demo_video = request.POST['demo_video']
                            if team.product_video:
                                team.product_video.delete()
                            team.app_video = demo_video
                        else:
                            demo_video = request.POST['demo_video']
                            if team.app_video:
                                team.app_video.delete()
                            team.product_video = demo_video

                    # synthesize a full file path; note that we included the filename
                    '''ppt_path_within_bucket = os.path.join(
                        file_path_bucket,
                        ppt.name
                    )
                    document_path_within_bucket = os.path.join(
                        file_path_bucket,
                        document.name
                    )
                    rs_paper_path_within_bucket = os.path.join(
                        file_path_bucket,
                        rs_paper.name
                    )
                    guide_form_path_within_bucket = os.path.join(
                        file_path_bucket,
                        guide_form.name
                    )'''

                    # if team.type == 'App Based':
                    #     if media_storage.exists(app_video_directory_within_bucket):
                    #         media_storage.delete(
                    #             app_video_directory_within_bucket)
                    # else:
                    #     if media_storage.exists(
                    #         product_video_directory_within_bucket
                    #     ):
                    #         media_storage.delete(
                    #             product_video_directory_within_bucket)

                    # if doc_storage.exists(ppt_path_within_bucket):
                    #     doc_storage.delete(ppt.name)
                    # if doc_storage.exists(document_path_within_bucket):
                    #     doc_storage.delete(document.name)
                    # if doc_storage.exists(rs_paper_path_within_bucket):
                    #     doc_storage.delete(rs_paper.name)
                    # if doc_storage.exists(guide_form_path_within_bucket):
                    #     doc_storage.delete(guide_form.name)

                    # doc_storage.save(file_path_within_bucket, ppt)
                    # file_url =

                    # team.ppt = doc_storage.save(file_path_within_bucket, ppt)
                    # team.document = doc_storage.save(
                    #     file_path_within_bucket, document)
                    # team.rs_paper = doc_storage.save(
                    #     file_path_within_bucket, rs_paper)
                    # team.guide_form = doc_storage.save(
                    #     file_path_within_bucket, guide_form)

                team.type = type
                team.save()
                # print('Type of project is: ', team.type)

                # auth.logout(request)
                return redirect('submitted')
            else:
                # POST else
                team = Team.objects.filter(teamID=user.username).get()
                context = {
                    'team': team,
                }
                return render(request, 'upload_docs/docs.html', context)
        else:
            messages.error(
                request, "Kindly wait until you are further instructed to upload documents")
            return redirect('team-dashboard')
    else:
        messages.error(request, "You're not logged in!")
        return redirect('login')


def profile(request):
    user = request.user
    if not user.is_authenticated:
        messages.error(request, "You're not Logged In!")
        return redirect('login')
    team = Team.objects.filter(teamID=user.username).get()
    if request.method == 'POST':
        project_name = request.POST['project_name']
        project_domain = request.POST['project_domain']
        project_description = request.POST['project_description']
        student_1_no = request.POST['student_1_no']

        if team.no_of_members == '2':
            student_2_no = request.POST['student_2_no']

            team = Team.objects.update(
                project_name=project_name,
                project_domain=project_domain,
                project_description=project_description,
                student_1_no=student_1_no,
                student_2_no=student_2_no,
            )
        else:
            team = Team.objects.update(
                project_name=project_name,
                project_domain=project_domain,
                project_description=project_description,
                student_1_no=student_1_no,
            )

        return redirect('upload')

    context = {
        'team': team
    }

    return render(request, 'dashboard/team_profile.html', context)


def custom_page_not_found_view(request, exception):
    return render(request, "errors/404.html", context={'exception': exception})


def my_custom_error_view(request):
    return render(request, 'errors/500.html')


def my_custom_permission_denied_view(request, exception):
    return render(request, 'errors/403.html', context={'exception': exception})


def my_custom_bad_request_view(request, exception):
    return render(request, 'errors/400.html', context={'exception': exception})


# conditional export


def export_to_excel(request):
    # apply filters as necessary
    queryset = Team.objects.filter(guide_approved=True).order_by('teamID')

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add column headings to the worksheet
    fields = Team._meta.get_fields()
    col = 1
    for field in fields:
        ws.cell(row=1, column=col, value=field.name.capitalize())
        col += 1

    # Add data to the worksheet
    row = 2
    for obj in queryset:
        col = 1
        for field in fields:
            value = getattr(obj, field.name)
            ws.cell(row=row, column=col, value=str(value))
            col += 1
        row += 1

    # Create an HTTP response with the Excel file as the content
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
    wb.save(response)

    return response
